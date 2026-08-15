"""Unit tests for parse_item_list, find_unmatched_skus, find_sku_churn, and
the catalog import's code-anchored rename detection.

parse_item_list parses the POS system's "item list" export (a different
shape than the stock statement — brand-header rows + item rows, no "Sub
Total"). find_unmatched_skus is the pure set-difference that flags any
currently-stocked SKU name absent from that catalog entirely. find_sku_churn
is the report-to-report diff that triggers the "go check your item list"
prompt; a rename is only paired off once db.import_item_catalog has actually
recorded that code's old name -> new name transition (there's no code in the
stock statement itself to resolve it from that side).
"""
import openpyxl
import pandas as pd
import pytest

from src.gowri_proj import db
from src.gowri_proj.analysis import find_sku_churn, find_unmatched_skus
from src.gowri_proj.parser import parse_item_list


def _write_item_list(path, rows, company="JANHAVI MEDICALS", location="BANGALORE", as_of="09/08/2026"):
    """rows: list of either ("brand", name) or ("item", code, product, packing,
    mrp, by_rate, tax_pct, hsn, long_name) tuples.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    sheet2 = wb.create_sheet("Sheet2")
    sheet2.append([company])
    sheet2.append([location])
    sheet2.append([f"Item List as on {as_of}"])
    sheet2.append(["Code", "Product", "Packing", "M.R.P.", "By.Rate", "Tax%", "Sp.Rate", "HSN", "Long Name"])
    for row in rows:
        if row[0] == "brand":
            sheet2.append([row[1]])
        else:
            _, code, product, packing, mrp, by_rate, tax_pct, hsn, long_name = row
            sheet2.append([code, product, packing, mrp, by_rate, tax_pct, None, hsn, long_name])
    wb.save(path)


def test_parses_items_and_tracks_brand_headers(tmp_path):
    path = tmp_path / "item_list.xlsx"
    _write_item_list(
        path,
        [
            ("brand", "CIPLA"),
            ("item", "C001", "ARKAMIN 0.1MG TAB", "10S", 20.0, 15.0, 12.0, "3004", "ARKAMIN 0.1MG TABLET"),
            ("item", "C002", "BUDAMATE 200 INHALER", "1S", 300.0, 250.0, 12.0, "3004", "BUDAMATE 200 INH"),
            ("brand", "GSK"),
            ("item", "G001", "AUGMENTIN 625 TAB", "10S", 200.0, 180.0, 12.0, "3004", "AUGMENTIN 625 DUO TAB"),
        ],
    )
    df, meta = parse_item_list(str(path))

    assert list(df["code"]) == ["C001", "C002", "G001"]
    assert list(df["brand"]) == ["CIPLA", "CIPLA", "GSK"]
    assert df.iloc[0]["product_name"] == "ARKAMIN 0.1MG TAB"
    assert df.iloc[0]["long_name"] == "ARKAMIN 0.1MG TABLET"
    assert df.iloc[0]["mrp"] == 20.0
    assert meta.company == "JANHAVI MEDICALS"
    assert meta.location == "BANGALORE"
    assert meta.as_of.isoformat() == "2026-08-09"


def test_all_codes_present_and_unique(tmp_path):
    path = tmp_path / "item_list.xlsx"
    _write_item_list(
        path,
        [
            ("brand", "CIPLA"),
            ("item", "C001", "A", "10S", 1, 1, 12, "3004", "A LONG"),
            ("item", "C002", "B", "10S", 1, 1, 12, "3004", "B LONG"),
        ],
    )
    df, _ = parse_item_list(str(path))
    assert df["code"].is_unique
    assert not df["code"].isna().any()


def test_raises_on_mutated_header(tmp_path):
    path = tmp_path / "item_list.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    sheet2 = wb.create_sheet("Sheet2")
    sheet2.append(["JANHAVI MEDICALS"])
    sheet2.append(["BANGALORE"])
    sheet2.append(["Item List as on 09/08/2026"])
    # "Product" moved — the export template changed shape.
    sheet2.append(["Code", "Packing", "Product", "M.R.P.", "By.Rate", "Tax%", "Sp.Rate", "HSN", "Long Name"])
    sheet2.append(["C001", "10S", "ARKAMIN", 20.0, 15.0, 12.0, None, "3004", "ARKAMIN LONG"])
    wb.save(path)

    with pytest.raises(ValueError, match="don't match the"):
        parse_item_list(str(path))


def test_empty_item_list_returns_empty_df(tmp_path):
    path = tmp_path / "item_list.xlsx"
    _write_item_list(path, [])
    df, _ = parse_item_list(str(path))
    assert df.empty
    assert list(df.columns) == [
        "code", "brand", "product_name", "packing", "mrp", "by_rate", "tax_pct", "hsn", "long_name",
    ]


def _enriched(rows):
    """rows: dicts with sku, brand, closing_stock, value."""
    df = pd.DataFrame(rows)
    if "brand" not in df.columns:
        df["brand"] = "SOME BRAND"
    for col in ("closing_stock", "value"):
        if col not in df.columns:
            df[col] = 0.0
    return df


def test_find_unmatched_skus_flags_names_absent_from_catalog():
    enriched = _enriched(
        [
            {"sku": "ARKAMIN 0.1MG TAB", "brand": "CIPLA", "closing_stock": 5, "value": 100.0},
            {"sku": "RENAMED PRODUCT XYZ", "brand": "CIPLA", "closing_stock": 3, "value": 50.0},
        ]
    )
    catalog_names = {"ARKAMIN 0.1MG TAB", "ARKAMIN 0.1MG TABLET"}
    result = find_unmatched_skus(enriched, catalog_names)
    assert result["total"] == 1
    assert [r["sku"] for r in result["items"]] == ["RENAMED PRODUCT XYZ"]


def test_find_unmatched_skus_matches_via_long_name_too():
    enriched = _enriched([{"sku": "AUGMENTIN 625 DUO TAB", "brand": "GSK", "closing_stock": 5, "value": 100.0}])
    # Only present as the catalog's long_name, not its product_name.
    catalog_names = {"AUGMENTIN 625 TAB", "AUGMENTIN 625 DUO TAB"}
    result = find_unmatched_skus(enriched, catalog_names)
    assert result == {"total": 0, "items": []}


def test_find_unmatched_skus_empty_catalog_flags_nothing():
    # No catalog imported yet — nothing to compare against, so nothing to flag.
    enriched = _enriched([{"sku": "ANYTHING", "brand": "X", "closing_stock": 1, "value": 1.0}])
    assert find_unmatched_skus(enriched, set()) == {"total": 0, "items": []}


def test_find_unmatched_skus_total_reflects_true_count_beyond_display_limit():
    # The display list is capped at `limit`, but "total" must still report
    # the true count — the UI headline ("N SKUs don't match...") would
    # otherwise silently understate the problem whenever it's large.
    enriched = _enriched(
        [{"sku": f"UNMATCHED {i}", "brand": "X", "closing_stock": 1, "value": float(i)} for i in range(5)]
    )
    result = find_unmatched_skus(enriched, catalog_names={"SOMETHING ELSE ENTIRELY"}, limit=2)
    assert result["total"] == 5
    assert len(result["items"]) == 2
    # Highest-value ones first.
    assert [r["sku"] for r in result["items"]] == ["UNMATCHED 4", "UNMATCHED 3"]


def _all_entries(reports: list[dict]) -> pd.DataFrame:
    """reports: list of {report_id, period_end, skus: [{sku, brand, closing_stock, value}]}.
    period_start defaults to period_end (these tests don't exercise the
    period_start tiebreak — see test_report_period_tiebreak.py for that).
    """
    rows = []
    for r in reports:
        period_start = r.get("period_start", r["period_end"])
        for sku_row in r["skus"]:
            rows.append(
                {"report_id": r["report_id"], "period_start": period_start, "period_end": r["period_end"], **sku_row}
            )
    df = pd.DataFrame(rows)
    df["period_start"] = pd.to_datetime(df["period_start"])
    df["period_end"] = pd.to_datetime(df["period_end"])
    return df


def test_find_sku_churn_needs_at_least_two_reports():
    entries = _all_entries(
        [{"report_id": 1, "period_end": "2026-07-31", "skus": [{"sku": "A", "brand": "X", "closing_stock": 1, "value": 1.0}]}]
    )
    result = find_sku_churn(entries)
    assert result == {
        "new_skus": [], "new_total": 0,
        "vanished_skus": [], "vanished_total": 0,
        "likely_renames": [], "renames_total": 0,
        "previous_period_end": None,
    }


def test_find_sku_churn_flags_new_and_vanished_names():
    entries = _all_entries(
        [
            {
                "report_id": 1, "period_end": "2026-06-30",
                "skus": [
                    {"sku": "STAYS", "brand": "X", "closing_stock": 1, "value": 1.0},
                    {"sku": "OLD NAME", "brand": "X", "closing_stock": 5, "value": 500.0},
                ],
            },
            {
                "report_id": 2, "period_end": "2026-07-31",
                "skus": [
                    {"sku": "STAYS", "brand": "X", "closing_stock": 1, "value": 1.0},
                    {"sku": "NEW NAME", "brand": "X", "closing_stock": 5, "value": 500.0},
                ],
            },
        ]
    )
    result = find_sku_churn(entries)
    assert result["new_total"] == 1
    assert [r["sku"] for r in result["new_skus"]] == ["NEW NAME"]
    assert result["vanished_total"] == 1
    assert [r["sku"] for r in result["vanished_skus"]] == ["OLD NAME"]
    assert result["previous_period_end"] == "2026-06-30"


def test_find_sku_churn_only_compares_the_latest_two_reports():
    # Churn from further back was already noticed (or wasn't) back then —
    # re-surfacing it forever on every visit would just be noise.
    entries = _all_entries(
        [
            {"report_id": 1, "period_end": "2026-05-31", "skus": [{"sku": "ANCIENT", "brand": "X", "closing_stock": 1, "value": 1.0}]},
            {"report_id": 2, "period_end": "2026-06-30", "skus": [{"sku": "STABLE", "brand": "X", "closing_stock": 1, "value": 1.0}]},
            {"report_id": 3, "period_end": "2026-07-31", "skus": [{"sku": "STABLE", "brand": "X", "closing_stock": 1, "value": 1.0}]},
        ]
    )
    result = find_sku_churn(entries)
    assert result["new_total"] == 0
    assert result["vanished_total"] == 0


def test_find_sku_churn_pairs_a_rename_confirmed_by_the_alias_map():
    # The alias map (db.get_name_change_map) says code C1 went from
    # "AMLOKIND 5MG TAB" to "AMLOKIND 5MG TAB (NON)" — a code-anchored fact,
    # not a guess — so this pairs off instead of counting as 1 new + 1
    # vanished.
    entries = _all_entries(
        [
            {"report_id": 1, "period_end": "2026-06-30", "skus": [{"sku": "AMLOKIND 5MG TAB", "brand": "MICRO", "closing_stock": 5, "value": 500.0}]},
            {"report_id": 2, "period_end": "2026-07-31", "skus": [{"sku": "AMLOKIND 5MG TAB (NON)", "brand": "MICRO", "closing_stock": 5, "value": 500.0}]},
        ]
    )
    alias_map = {"AMLOKIND 5MG TAB": "AMLOKIND 5MG TAB (NON)"}
    result = find_sku_churn(entries, alias_map)
    assert result["new_total"] == 0
    assert result["vanished_total"] == 0
    assert result["renames_total"] == 1
    assert result["likely_renames"] == [
        {"brand": "MICRO", "old_name": "AMLOKIND 5MG TAB", "new_name": "AMLOKIND 5MG TAB (NON)", "closing_stock": 5.0, "value": 500.0}
    ]


def test_find_sku_churn_does_not_pair_without_a_confirming_alias():
    # No item list uploaded yet (alias_map is empty) — even an obviously
    # cosmetic difference like "10G" -> "10GM" stays an unresolved new +
    # vanished pair. There's no code in the stock statement itself to
    # resolve it from; only a fresh catalog import can.
    entries = _all_entries(
        [
            {"report_id": 1, "period_end": "2026-06-30", "skus": [{"sku": "XYZ 10G", "brand": "MICRO", "closing_stock": 5, "value": 500.0}]},
            {"report_id": 2, "period_end": "2026-07-31", "skus": [{"sku": "XYZ 10GM", "brand": "MICRO", "closing_stock": 5, "value": 500.0}]},
        ]
    )
    result = find_sku_churn(entries)
    assert result["renames_total"] == 0
    assert result["new_total"] == 1
    assert result["vanished_total"] == 1


def test_find_sku_churn_does_not_pair_a_different_strength_even_with_an_unrelated_alias():
    # A real product swap (different strength) shouldn't be swallowed just
    # because the alias map happens to contain unrelated entries — the old
    # name has to be a key that maps to exactly this new name, nothing looser.
    entries = _all_entries(
        [
            {"report_id": 1, "period_end": "2026-06-30", "skus": [{"sku": "DOLO 650", "brand": "MICRO", "closing_stock": 5, "value": 500.0}]},
            {"report_id": 2, "period_end": "2026-07-31", "skus": [{"sku": "DOLO 350", "brand": "MICRO", "closing_stock": 5, "value": 500.0}]},
        ]
    )
    alias_map = {"SOMETHING ELSE": "SOMETHING ELSE V2"}
    result = find_sku_churn(entries, alias_map)
    assert result["renames_total"] == 0
    assert result["new_total"] == 1
    assert result["vanished_total"] == 1


def test_find_sku_churn_ignores_an_alias_whose_target_isnt_in_the_new_report():
    # The catalog recorded a rename, but the "new" name it points to isn't
    # actually present in this report pair (e.g. the alias is stale, or
    # belongs to a different period) — shouldn't fabricate a pairing.
    entries = _all_entries(
        [
            {
                "report_id": 1, "period_end": "2026-06-30",
                "skus": [{"sku": "OLD NAME", "brand": "X", "closing_stock": 5, "value": 500.0}],
            },
            {
                "report_id": 2, "period_end": "2026-07-31",
                "skus": [{"sku": "UNRELATED STABLE SKU", "brand": "X", "closing_stock": 1, "value": 1.0}],
            },
        ]
    )
    alias_map = {"OLD NAME": "NEW NAME THAT NEVER SHOWS UP"}
    result = find_sku_churn(entries, alias_map)
    assert result["renames_total"] == 0
    assert result["vanished_total"] == 1


def _catalog_df(rows):
    """rows: dicts with code, brand, product_name, packing, mrp, by_rate,
    tax_pct, hsn, long_name."""
    cols = ["code", "brand", "product_name", "packing", "mrp", "by_rate", "tax_pct", "hsn", "long_name"]
    return pd.DataFrame(rows)[cols]


def _catalog_row(code, product_name, long_name, brand="MICRO"):
    return {
        "code": code, "brand": brand, "product_name": product_name, "packing": "1S",
        "mrp": 1.0, "by_rate": 1.0, "tax_pct": 12.0, "hsn": "3004", "long_name": long_name,
    }


def test_import_item_catalog_logs_a_rename_when_a_codes_name_changes(tmp_path):
    with db.connect(str(tmp_path / "test.db")) as conn:
        db.import_item_catalog(conn, _catalog_df([_catalog_row("C1", "XYZ 10G", "XYZ 10G TUBE")]))
        db.import_item_catalog(conn, _catalog_df([_catalog_row("C1", "XYZ 10GM", "XYZ 10GM TUBE")]))
        alias_map = db.get_name_change_map(conn)
    assert alias_map["XYZ 10G"] == "XYZ 10GM"
    assert alias_map["XYZ 10G TUBE"] == "XYZ 10GM TUBE"


def test_import_item_catalog_logs_nothing_when_names_are_unchanged(tmp_path):
    with db.connect(str(tmp_path / "test.db")) as conn:
        db.import_item_catalog(conn, _catalog_df([_catalog_row("C1", "XYZ 10G", "XYZ 10G TUBE")]))
        db.import_item_catalog(conn, _catalog_df([_catalog_row("C1", "XYZ 10G", "XYZ 10G TUBE")]))
        alias_map = db.get_name_change_map(conn)
    assert alias_map == {}


def test_import_item_catalog_logs_nothing_for_a_brand_new_code(tmp_path):
    # A code that's never been seen before has no "old" name to diff
    # against — it's a genuinely new catalog entry, not a rename.
    with db.connect(str(tmp_path / "test.db")) as conn:
        db.import_item_catalog(conn, _catalog_df([_catalog_row("C1", "XYZ 10G", "XYZ 10G TUBE")]))
        db.import_item_catalog(
            conn,
            _catalog_df([_catalog_row("C1", "XYZ 10G", "XYZ 10G TUBE"), _catalog_row("C2", "NEW ITEM", "NEW ITEM LONG")]),
        )
        alias_map = db.get_name_change_map(conn)
    assert alias_map == {}


def test_import_item_catalog_replaces_the_whole_table(tmp_path):
    with db.connect(str(tmp_path / "test.db")) as conn:
        db.import_item_catalog(conn, _catalog_df([_catalog_row("C1", "A", "A LONG")]))
        count = db.import_item_catalog(conn, _catalog_df([_catalog_row("C2", "B", "B LONG")]))
        meta = db.get_item_catalog_meta(conn)
    assert count == 1
    assert meta["item_count"] == 1
